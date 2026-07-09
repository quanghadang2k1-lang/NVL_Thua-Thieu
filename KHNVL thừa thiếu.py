def generate_excel(allocated_df, processed_df=None, pivot=None, merged_inventory=None, summary_df=None):
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    import io
    import pandas as pd

    allocated_df = allocated_df.copy()
    if 'Remaining_Stock' in allocated_df.columns:
        allocated_df = allocated_df.rename(columns={'Remaining_Stock': 'Thừa thiếu'})

    # Reorder columns to group by metric instead of product
    all_cols = list(allocated_df.columns)
    fixed_start = [c for c in ['Level Group', 'Allocation Pool', 'Filter VNPT MAN P/N', 'Description', 'Popularity'] if c in all_cols]
    std_cols_name = [c for c in all_cols if ' - Standard Qty' in c]
    kh_cols_name = [c for c in all_cols if ' - SL theo KH' in c]
    alloc_cols_name = [c for c in all_cols if ' - SL sau phân bổ kho' in c]

    # Preserve any remaining columns at the end
    grouped_set = set(fixed_start + std_cols_name + kh_cols_name + alloc_cols_name)
    end_cols = [c for c in all_cols if c not in grouped_set]

    new_order = fixed_start + std_cols_name + kh_cols_name + alloc_cols_name + end_cols
    allocated_df = allocated_df[new_order]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        allocated_df.to_excel(writer, sheet_name='Allocated', index=False)

        if processed_df is not None:
            processed_df.to_excel(writer, sheet_name='BOM Result', index=False)
        if pivot is not None:
            pivot.to_excel(writer, sheet_name='Pivot', index=False)
        if merged_inventory is not None:
            merged_inventory.to_excel(writer, sheet_name='Inventory', index=False)
        if summary_df is not None:
            summary_df.to_excel(writer, sheet_name='KHSX', index=False)

        worksheet = writer.sheets['Allocated']
        alloc_pool_col_idx = None
        thua_thieu_col_idx = None
        max_col = worksheet.max_column

        std_cols = []
        kh_cols = []
        alloc_cols = []

        for col_idx, col_name in enumerate(allocated_df.columns, 1):
            col_str = str(col_name)
            if col_str == 'Allocation Pool':
                alloc_pool_col_idx = col_idx
            elif col_str == 'Thừa thiếu':
                thua_thieu_col_idx = col_idx

            if ' - Standard Qty' in col_str:
                std_cols.append(col_idx)
            elif ' - SL theo KH' in col_str:
                kh_cols.append(col_idx)
            elif ' - SL sau phân bổ kho' in col_str:
                alloc_cols.append(col_idx)

        fill_all = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        fill_std = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        fill_kh = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        fill_alloc = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        bold_font = Font(bold=True)
        thin_side = Side(style='thin')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = fill_all
            cell.font = bold_font
            cell.border = thin_border

        for col_idx in std_cols:
            worksheet.cell(row=1, column=col_idx).fill = fill_std
        for col_idx in kh_cols:
            worksheet.cell(row=1, column=col_idx).fill = fill_kh
        for col_idx in alloc_cols:
            worksheet.cell(row=1, column=col_idx).fill = fill_alloc

        sep_cols = []
        if std_cols:
            sep_cols.append(min(std_cols) - 1)
            sep_cols.append(max(std_cols))
        if kh_cols:
            sep_cols.append(max(kh_cols))
        if alloc_cols:
            sep_cols.append(max(alloc_cols))

        medium_side = Side(style='medium')

        for r in range(1, len(allocated_df) + 2):
            for c in sep_cols:
                if c > 0:
                    cell = worksheet.cell(row=r, column=c)
                    cell.border = Border(left=cell.border.left, right=medium_side, top=cell.border.top, bottom=cell.border.bottom)

        # --- Apply font color for negative remaining stock ---
        if thua_thieu_col_idx:
            red_font = Font(color="FF0000", bold=True)
            for r in range(2, len(allocated_df) + 2):
                thua_thieu_val = worksheet.cell(row=r, column=thua_thieu_col_idx).value
                if isinstance(thua_thieu_val, (int, float)) and thua_thieu_val < -0.0001:
                    for c in alloc_cols:
                        cell = worksheet.cell(row=r, column=c)
                        if isinstance(cell.value, (int, float)) and cell.value > 0:
                            # Check if the allocated value is less than the requested value
                            col_name = allocated_df.columns[c - 1]
                            kh_col_name = col_name.replace(' - SL sau phân bổ kho', ' - SL theo KH')
                            kh_val = allocated_df.iloc[r - 2].get(kh_col_name, 0)
                            if abs(cell.value - kh_val) > 0.0001:
                                cell.font = red_font

        if alloc_pool_col_idx:
            center_alignment = Alignment(horizontal='center', vertical='center')
            medium_bottom_border = Side(style='medium')

            for r in range(1, len(allocated_df) + 2):
                worksheet.cell(row=r, column=alloc_pool_col_idx).alignment = center_alignment

            start_row = 2
            for i in range(1, len(allocated_df)):
                val = allocated_df['Allocation Pool'].iloc[i]
                current_excel_row = i + 2
                if val != '':
                    prev_row = current_excel_row - 1
                    if prev_row > start_row:
                        worksheet.merge_cells(start_row=start_row, start_column=alloc_pool_col_idx, end_row=prev_row, end_column=alloc_pool_col_idx)
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=prev_row, column=col)
                        cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=medium_bottom_border)
                    start_row = current_excel_row

            last_row = len(allocated_df) + 1
            if last_row > start_row:
                worksheet.merge_cells(start_row=start_row, start_column=alloc_pool_col_idx, end_row=last_row, end_column=alloc_pool_col_idx)
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=last_row, column=col)
                cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=medium_bottom_border)

    output.seek(0)
    return output
